import sys
import math
import pprint
import csv
from openpyxl import load_workbook
from argparse import ArgumentParser

def mean(numbers):
    # print(numbers)
    s = float(sum(numbers))
    # print(f"SUM: {s}")
    return s / max(len(numbers), 1)

def trunc_float(f, n):
    return math.floor(f * 10 ** n) / 10 ** n

def calc_handicap(n):
    return math.floor(0.96 * n)

# Takes a sorted list of scores, lowest to highest
# Returns the average of the correct number of low scores
def avg_course_index(lowest_scores):
    if len(lowest_scores) < 7:
        n = 1
    elif len(lowest_scores) < 9:
        n = 2
    elif len(lowest_scores) < 10:
        n = 3
    elif len(lowest_scores) < 12:
        n = 4
    elif len(lowest_scores) < 14:
        n = 5
    elif len(lowest_scores) < 16:
        n = 6
    elif len(lowest_scores) == 17:
        n = 7
    elif len(lowest_scores) == 18:
        n = 8
    elif len(lowest_scores) == 19:
        n = 9
    else:
        n = 10
    return trunc_float(mean(lowest_scores[0:n]), 1)

parser = ArgumentParser()
parser.add_argument("spreadsheet")
parser.add_argument("-o", metavar="output", help="If provided, output to this file. Otherwise, overwrite the original")
args = parser.parse_args()

# Run formulas the first time to get things like diff
wb = load_workbook(filename=args.spreadsheet,
    data_only=True)
players_sheet = wb['Players']

players_data = players_sheet.rows

player = None
scores = []
data_row = False
header_row = False
score_rows = False
final_players = {}
pp = pprint.PrettyPrinter(indent=4)
row_replacements = []
player_row_starts = []
try:
    for idx, r in enumerate(players_data):
        # print(f"Row: {[x.value for x in r]}")
        if header_row:
            # print("Header row")
            header_row = False
            score_rows = True
        elif data_row:
            # Pull in player data
            player_data = [x.value for x in r[4:9]]
            player_data = list(["{}{}".format(x.column, x.row) for x in r[4:9]])
            # print(f"Player data: {player_data}")
            data_row = False
            header_row = True
        elif r[4].value == "Handicap":
            # We have a new player
            player = r[1].value
            print(f"Processing player: {player}")
            # print([x.value for x in r])
            data_row = True
        elif score_rows and (r[0].value == 21 or (r[1] is not None and r[1].value is None)):
            # If we have a final value, add it for consideration
            if r[0].value == 21 and r[1].value is not None:
                # print(f"Row: {[x.value for x in r]}")
                # print(f"Adding last score")
                row_idx = idx+1
                scores.append((row_idx, [x.value for x in r]))
            # print(f"Finished scores for {player}")
            # First, if there are more than 20 scores, find the
            # 20 most recent
            player_row_starts.append(scores[0][0])
            if len(scores) > 20:
                # print(len(scores))
                sorted_by_date = sorted(scores, key=lambda x: x[1][1])
                # pp.pprint([x[1][1] for x in sorted_by_date])
                # print(f"Will replace row {sorted_by_date[0][0]} with {scores[20][0]}")
                row_replacements.append((scores[20][0], sorted_by_date[0][0]))
                scores = [x[1] for x in sorted_by_date][1:]
            else:
                # Knock off indexes
                scores = [x[1] for x in scores]
            # Calculate new values
            sorted_by_score = sorted(scores, key=lambda x: x[4])
            differentials = [x[8] for x in sorted_by_score]
            # print(f"Diffs:")
            # pp.pprint(differentials)
            ai = avg_course_index(differentials)
            # print(f"Avg Course Index: {ai}")
            hc = calc_handicap(ai)
            # print(f"Handicap: {hc}")
            # TODO: Save off new values for player
            final_players[player] = {
                player_data[0]: hc,
                player_data[1]: ai
            }
            # Clear data for next player
            player = None
            score_rows = False
            scores = []
        elif score_rows:
            # Scores
            row_idx = idx+1
            # print(f"Scores {row_idx}: {[x.value for x in r]}")
            scores.append((row_idx, [x.value for x in r]))
        else:
            pass
except:
    import traceback
    print("Error encountered:")
    traceback.print_exc()
    print("\n\n!!! Make sure you've opened and saved the input XLSX in Excel since last time")
    sys.exit(1)
wb.close()
wb = load_workbook(filename=args.spreadsheet)
players_sheet = wb['Players']
# Update player data
for player, data in final_players.items():
    # print(player, data)
    for k, v in data.items():
        players_sheet[k] = v
# Remove rows and renumber
for rd in row_replacements:
    print(f"Replacing row {rd[1]} with {rd[0]}")
    for l in ("B", "C", "D", "E", "F"):
        players_sheet[l + str(rd[1])] = players_sheet[l + str(rd[0])].value
        players_sheet[l + str(rd[0])] = ""
    # Shift all number rows
    # nprs = []
    # for f in player_row_starts:
    #     if f > rd:
    #         nprs.append(f-1)
    #     else:
    #         nprs.append(f)
    # player_row_starts = nprs

# Update player row numbering
for f in player_row_starts:
    for i in range(f, f+21):
        # print(f"Updating A{i} to {i-f+1}")
        players_sheet[f"A{i}"] = i-f+1

# Save out spreadsheet
if args.o is None:
    # Overwrite
    print(f"Writing to {args.spreadsheet}")
    wb.save(args.spreadsheet)
else:
    print(f"Writing to {args.o}")
    wb.save(args.o)